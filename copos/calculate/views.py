from django.shortcuts import render,redirect,get_object_or_404
from django.contrib.auth import authenticate, login, logout
from .forms import RegisterForm
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import Course, CO, PO, COPOMapping, COAttainment, StudentMark,Student
from django.contrib.auth.models import User
from openpyxl import load_workbook
from collections import defaultdict

# Create your views here.
def home(request):
    return render(request, 'home.html')


def register(request):
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password1')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                return render(request, 'home.html', {'message': 'Registration successful!'})
    else:
        form = RegisterForm()
    return render(request, 'register.html', {'form': form})



def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return render(request, 'home.html', {
                'message': 'Login successful!',
                'user': user
            })
        else:
            return render(request, 'login.html', {
                'error': 'Invalid credentials'
            })
    return render(request, 'login.html')


def logout_view(request):
    logout(request)
    return redirect('login') 

def add_course(request):
    if request.method == 'POST':
        course_code = request.POST.get('course_code')
        course_name = request.POST.get('course_name')
        user_id = request.POST.get('user')  # 👈 get from form

        try:
            user = User.objects.get(id=user_id)
            Course.objects.create(
                code=course_code,
                name=course_name,
                user=user  # ✅ set user from form
            )
            return redirect('add_course')  # or show success page
        except User.DoesNotExist:
            return render(request, 'add_course.html', {
                'error': 'Invalid user selected.'
            })

    users = User.objects.all()
    return render(request, 'add_course.html', {
        'users': users,
        'user': request.user
    })


def courses(request):
    if request.user.is_authenticated:
        courses = Course.objects.filter(user=request.user)
        return render(request, 'courses.html', {'courses': courses})
    else:
        return redirect('login')  # Redirect to login if not authenticated
    
@login_required
def add_co(request):
    if request.method == 'POST':
        course_id = request.POST.get('course')
        co_number = request.POST.get('co_number')
        co_description = request.POST.get('description')
        max_marks = request.POST.get('max_marks')  # <-- get from form

        try:
            course = Course.objects.get(id=course_id, user=request.user)

            if CO.objects.filter(course=course, number=co_number).exists():
                messages.warning(request, "This CO already exists for the selected course.")
            else:
                CO.objects.create(
                    course=course,
                    number=co_number,
                    description=co_description,
                    max_score=max_marks  # <-- save to db
                )
                messages.success(request, "Course Outcome added successfully!")
            return redirect('add_co')

        except Course.DoesNotExist:
            messages.error(request, "Invalid course selected.")

    courses = Course.objects.filter(user=request.user)
    return render(request, 'add_co.html', {'courses': courses})

def add_po(request):
    if request.method == 'POST':
        po_number = request.POST.get('po_number')
        po_description = request.POST.get('po_description')

        PO.objects.create(
            number=po_number,
            description=po_description
        )
        return redirect('add_po')  # or show success page

    return render(request, 'add_po.html')


def add_mapping(request):
    if request.method == 'POST':
        processed = False
        course_id = request.POST.get('course')
        if not course_id:
            messages.error(request, "Course selection is required for mapping.")
            return redirect('add_mapping')

        # Excel upload
        excel_file = request.FILES.get('excel_file')
        if excel_file:
            try:
                wb = load_workbook(excel_file)
                sheet = wb.active

                # Read PO headers from first row (skip first column)
                po_headers = [str(cell.value).strip().upper() for cell in sheet[1][1:]]

                # Prepare PO and CO objects for fast lookup
                po_map = {po.number.upper(): po for po in PO.objects.all()}
                co_map = {co.number.upper(): co for co in CO.objects.filter(course_id=course_id)}

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    co_key = str(row[0]).strip().upper()
                    co = co_map.get(co_key)

                    if not co:
                        continue

                    for i, level in enumerate(row[1:]):
                        po_key = po_headers[i]
                        po = po_map.get(po_key)

                        if po and level:
                            COPOMapping.objects.update_or_create(
                                co=co,
                                po=po,
                                defaults={'level': int(level)}
                            )
                messages.success(request, "CO-PO mappings updated from Excel file!")
                processed = True

            except Exception as e:
                messages.error(request, f"Failed to process Excel file: {e}")
                processed = True

        # Manual form mapping (unchanged)
        co_id = request.POST.get('co')
        po_id = request.POST.get('po')
        level = request.POST.get('level')
        if co_id and po_id and level:
            try:
                co = CO.objects.get(id=co_id, course_id=course_id)
                po = PO.objects.get(id=po_id)
                COPOMapping.objects.create(
                    co=co,
                    po=po,
                    level=level
                )
                messages.success(request, "CO-PO mapping added successfully!")
                processed = True
            except (CO.DoesNotExist, PO.DoesNotExist):
                messages.error(request, "Invalid CO or PO selected.")
                processed = True

        if processed:
            return redirect('add_mapping')
        else:
            messages.warning(request, "No mapping data provided.")

    cos = CO.objects.filter(course__user=request.user)
    pos = PO.objects.all()
    courses = Course.objects.filter(user=request.user)
    return render(request, 'mapping.html', {'cos': cos, 'pos': pos, 'courses': courses})


def upload_marks(request):
    courses = Course.objects.all()

    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        course_id = request.POST.get('course_id')

        try:
            course = Course.objects.get(id=course_id)
            cos = {co.number.strip().upper(): co for co in CO.objects.filter(course=course)}
        except Course.DoesNotExist:
            messages.error(request, "Invalid course selected.")
            return redirect('upload_marks')

        try:
            wb = load_workbook(excel_file)
            sheet = wb.active

            # ✅ Skip merged evaluation row (1st row), read header from row 2
            header = [str(cell.value).strip().upper() for cell in sheet[2]]
            co_headers = header[2:]  # Skip Roll No and Name
        except Exception as e:
            messages.error(request, f"Failed to read Excel file: {e}")
            return redirect('upload_marks')

        created_count = 0
        updated_count = 0

        # ✅ Start reading from row 3
        for row in sheet.iter_rows(min_row=3, values_only=True):
            if not row or row[0] is None:
                continue

            roll_no, name, *marks = row
            roll_no = str(roll_no).strip()
            name = str(name).strip()

            if not roll_no or not name:
                continue

            student, _ = Student.objects.get_or_create(
                roll_number=roll_no,
                defaults={'name': name}
            )
            if student.name != name:
                student.name = name
                student.save()

            # ✅ 1. Sum all marks by CO number
            co_marks_sum = defaultdict(float)
            for i, co_name in enumerate(co_headers):
                co_key = co_name.strip().upper()
                if i < len(marks) and marks[i] is not None:
                    try:
                        co_marks_sum[co_key] += float(marks[i])
                    except ValueError:
                        continue

            # ✅ 2. Save summed marks for each CO
            for co_key, total_obtained in co_marks_sum.items():
                co = cos.get(co_key)
                if co:
                    obj, created = StudentMark.objects.update_or_create(
                        course=course,
                        co=co,
                        student=student,
                        defaults={
                            'obtained_marks': total_obtained,
                            'total_marks': co.max_score  # Optional: You can update this logic
                        }
                    )
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1

        messages.success(request, f"{created_count} new records, {updated_count} updated.")
        return redirect('upload_marks')

    return render(request, 'upload_marks.html', {'courses': courses})

@login_required
def co_attainment_view(request):
    selected_course_id = request.GET.get('course_id')
    courses = Course.objects.filter(user=request.user)
    attainment_data = []
    selected_course = None
    overall_level_avg = 0

    if selected_course_id:
        selected_course = get_object_or_404(Course, id=selected_course_id, user=request.user)
        cos = CO.objects.filter(course=selected_course)

        total_level = 0
        co_count = 0

        for co in cos:
            student_marks = StudentMark.objects.filter(course=selected_course, co=co)
            levels = []

            for mark in student_marks:
                if mark.total_marks > 0:
                    percent = (mark.obtained_marks / mark.total_marks) * 100
                    if percent >= 60:
                        levels.append(3)
                    elif percent > 40:
                        levels.append(2)
                    else:
                        levels.append(1)

            if levels:
                avg_level = round(sum(levels) / len(levels), 2)
            else:
                avg_level = 0

            total_level += avg_level
            co_count += 1

            # ✅ Save only average level
            COAttainment.objects.update_or_create(
                course=selected_course,
                co=co,
                defaults={'level_avg': avg_level}
            )

            attainment_data.append({
                'co_number': co.number,
                'co_description': co.description,
                'level': avg_level
            })

        if co_count > 0:
            overall_level_avg = round(total_level / co_count, 2)

    return render(request, 'co_attainment.html', {
        'courses': courses,
        'selected_course': selected_course,
        'attainment_data': attainment_data,
        'average_level': overall_level_avg
    })

def calculate_po_attainment(request):
    selected_course_id = request.GET.get('course_id')
    courses = Course.objects.filter(user=request.user)
    po_scores = {}
    selected_course = None

    if selected_course_id:
        selected_course = get_object_or_404(Course, id=selected_course_id)
        pos = PO.objects.all()
        cos = CO.objects.filter(course=selected_course)

        for po in pos:
            total_score = 0
            total_weight = 0

            for co in cos:
                try:
                    mapping = COPOMapping.objects.get(co=co, po=po)
                    attainment = COAttainment.objects.get(co=co, course=selected_course).level_avg
                    total_score += attainment * mapping.level
                    total_weight += mapping.level
                except (COPOMapping.DoesNotExist, COAttainment.DoesNotExist):
                    continue

            po_scores[po.number] = round(total_score / total_weight, 2) if total_weight else 0

    return render(request, 'po_attainment.html', {
        'courses': courses,
        'selected_course': selected_course,
        'po_scores': po_scores
    })
